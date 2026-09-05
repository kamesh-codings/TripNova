#!/usr/bin/env python3
"""
=============================================================================
TripNova Tourism Dataset Synchronization Engine
=============================================================================
Unifies locations and places from:
1. backend/database/datasets/locations_tn.csv (38 Tamil Nadu districts)
2. backend/database/datasets/locations_kl.csv (14 Kerala districts)
3. backend/database/datasets/places_1.csv (71 places)
4. backend/database/datasets/places_2.csv (200 places)
5. backend/database/datasets/india_tourism_dataset-8.xlsx:
   - 'Dataset' (57 destinations across all 28 states + 8 UTs)
   - 'State_District_Spots' (235 curated spots across all Indian states)
   - 'Uploaded_Places_Extended' (1,019 places across India)
   - 'Uploaded_Top_Places_Kaggle' (325 top places)

Standardizes:
- Common Location IDs: 'loc-[state_code]-[district_slug]'
- Common Place IDs: 'plc-[place_slug]'
- Normalized Categories, Ratings, Entry Fees, Coordinates, Seasons, Durations,
  Transport, Nearby Hotels, and Dining hints.

Updates:
- backend/database/tripnova.db (Live SQLite database)
- backend/database/seed.sql (MySQL bootstrap and seeding file)
=============================================================================
"""

import os
import sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
import re
import csv
import sqlite3
import openpyxl
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATASETS_DIR = os.path.join(BASE_DIR, 'database', 'datasets')
DB_PATH = os.path.join(BASE_DIR, 'database', 'tripnova.db')
SEED_SQL_PATH = os.path.join(BASE_DIR, 'database', 'seed.sql')

STATE_CODE_MAP = {
    'andaman & nicobar islands': 'an',
    'andaman and nicobar islands': 'an',
    'andhra pradesh': 'ap',
    'arunachal pradesh': 'ar',
    'assam': 'as',
    'bihar': 'br',
    'chandigarh': 'ch',
    'chhattisgarh': 'cg',
    'dadra and nagar haveli and daman and diu': 'dn',
    'daman and diu': 'dn',
    'delhi': 'dl',
    'goa': 'ga',
    'gujarat': 'gj',
    'gujrat': 'gj',
    'haryana': 'hr',
    'himachal pradesh': 'hp',
    'jammu & kashmir': 'jk',
    'jammu and kashmir': 'jk',
    'ladakh': 'la',
    'lakshadweep': 'ld',
    'madhya pradesh': 'mp',
    'maharashtra': 'mh',
    'manipur': 'mn',
    'meghalaya': 'ml',
    'mizoram': 'mz',
    'nagaland': 'nl',
    'odisha': 'od',
    'puducherry': 'py',
    'pondicherry': 'py',
    'punjab': 'pb',
    'rajasthan': 'rj',
    'sikkim': 'sk',
    'tamil nadu': 'tn',
    'telangana': 'ts',
    'tripura': 'tr',
    'uttar pradesh': 'up',
    'uttarakhand': 'uk',
    'west bengal': 'wb'
}

STATE_REGION_MAP = {
    'tamil nadu': 'Southern',
    'kerala': 'Southern',
    'karnataka': 'Southern',
    'andhra pradesh': 'Southern',
    'telangana': 'Southern',
    'puducherry': 'Southern',
    'delhi': 'Northern',
    'punjab': 'Northern',
    'haryana': 'Northern',
    'himachal pradesh': 'Northern',
    'jammu & kashmir': 'Northern',
    'jammu and kashmir': 'Northern',
    'ladakh': 'Northern',
    'uttar pradesh': 'Northern',
    'uttarakhand': 'Northern',
    'chandigarh': 'Northern',
    'rajasthan': 'Western',
    'gujarat': 'Western',
    'maharashtra': 'Western',
    'goa': 'Western',
    'dadra and nagar haveli and daman and diu': 'Western',
    'daman and diu': 'Western',
    'madhya pradesh': 'Central',
    'chhattisgarh': 'Central',
    'bihar': 'Eastern',
    'jharkhand': 'Eastern',
    'odisha': 'Eastern',
    'west bengal': 'Eastern',
    'assam': 'North Eastern',
    'arunachal pradesh': 'North Eastern',
    'manipur': 'North Eastern',
    'meghalaya': 'North Eastern',
    'mizoram': 'North Eastern',
    'nagaland': 'North Eastern',
    'sikkim': 'North Eastern',
    'tripura': 'North Eastern',
    'andaman & nicobar islands': 'Islands',
    'andaman and nicobar islands': 'Islands',
    'lakshadweep': 'Islands'
}

def slugify(text, max_len=36):
    text = re.sub(r'[^\w\s-]', '', str(text or '')).strip().lower()
    slug = re.sub(r'[-\s]+', '-', text)
    return slug[:max_len].strip('-')

def normalize_state_name(state_raw):
    s = str(state_raw or '').strip()
    s_low = s.lower()
    if 'gujrat' in s_low: return 'Gujarat'
    if 'karanataka' in s_low: return 'Karnataka'
    if 'andaman' in s_low: return 'Andaman & Nicobar Islands'
    if 'jammu' in s_low: return 'Jammu & Kashmir'
    if 'daman' in s_low or 'dadra' in s_low: return 'Dadra and Nagar Haveli and Daman and Diu'
    if 'pondicherry' in s_low: return 'Puducherry'
    words = [w.capitalize() if w.lower() not in ['and', '&', 'of', 'the'] else w.lower() for w in s.split()]
    return ' '.join(words)

def normalize_category(cat_raw, name_raw=''):
    s = f"{cat_raw or ''} {name_raw or ''}".lower()
    if any(k in s for k in ['temple', 'religious', 'spiritual', 'church', 'mosque', 'gurudwara', 'basilica', 'dargah', 'pilgrimage', 'shrine', 'ashram', 'matha']):
        return 'religious'
    if any(k in s for k in ['beach', 'sea', 'coast', 'cove', 'promenade']):
        return 'beach'
    if any(k in s for k in ['museum', 'planetarium', 'gallery', 'memorial museum', 'science center']):
        return 'museum'
    if any(k in s for k in ['wildlife', 'national park', 'sanctuary', 'tiger reserve', 'safari', 'zoo', 'biosphere']):
        return 'wildlife'
    if any(k in s for k in ['hill', 'peak', 'ridge', 'valley', 'pass', 'viewpoint', 'plateau', 'ghat']):
        return 'hill_station'
    if any(k in s for k in ['falls', 'waterfall', 'lake', 'dam', 'caves', 'cavern', 'river', 'forest', 'garden', 'park', 'nature']):
        return 'nature'
    if any(k in s for k in ['fort', 'palace', 'tomb', 'monument', 'ruins', 'ancient', 'historical', 'anicut', 'stepwell', 'chhatri']):
        return 'historical'
    if any(k in s for k in ['heritage', 'unesco', 'haveli', 'village', 'mahal', 'memorial']):
        return 'heritage'
    if any(k in s for k in ['adventure', 'trek', 'rafting', 'paragliding', 'water sports', 'safari']):
        return 'adventure'
    if any(k in s for k in ['market', 'bazaar', 'shopping', 'silk', 'handloom', 'mall']):
        return 'shopping'
    if any(k in s for k in ['food', 'cuisine', 'restaurant', 'dining', 'mess']):
        return 'food_dining'
    if any(k in s for k in ['cultural', 'arts', 'dance', 'theatre', 'craft']):
        return 'cultural'
    return 'other'

def parse_rating(val):
    if val is None: return 4.5
    if isinstance(val, (int, float)):
        return round(min(5.0, max(1.0, float(val))), 2)
    s = str(val)
    m = re.findall(r'(\d+\.?\d*)', s)
    if m:
        nums = [float(x) for x in m if float(x) <= 5.0]
        if nums:
            return round(sum(nums) / len(nums), 2)
    return 4.5

def parse_entry_fee(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).lower()
    if 'free' in s: return 0.0
    m = re.findall(r'(?:rs\.?|inr|\u20b9)?\s*(\d+)', s)
    if m: return float(m[0])
    return 0.0

def load_locations():
    locations = {}
    loc_id_map = {}

    # 1. Load Tamil Nadu CSV
    path_tn = os.path.join(DATASETS_DIR, 'locations_tn.csv')
    if os.path.exists(path_tn):
        with open(path_tn, 'r', encoding='utf-8', errors='ignore') as f:
            for r in csv.DictReader(f):
                loc_id = r['id'].strip()
                name = r['name'].strip()
                state = normalize_state_name(r['state'])
                key = (name.lower(), state.lower())
                loc_obj = {
                    'id': loc_id,
                    'name': name,
                    'state': state,
                    'country': r.get('country', 'India').strip(),
                    'currency_code': r.get('currency_code', 'INR').strip(),
                    'description': r.get('description', '').strip(),
                    'latitude': float(r['latitude']) if r.get('latitude') else None,
                    'longitude': float(r['longitude']) if r.get('longitude') else None,
                    'region': r.get('region', 'Southern').strip(),
                    'created_at': r.get('created_at', '2026-09-04 00:00:00'),
                    'updated_at': r.get('updated_at', '2026-09-04 00:00:00')
                }
                locations[key] = loc_obj
                loc_id_map[loc_id] = loc_obj

    # 2. Load Kerala CSV
    path_kl = os.path.join(DATASETS_DIR, 'locations_kl.csv')
    if os.path.exists(path_kl):
        with open(path_kl, 'r', encoding='utf-8', errors='ignore') as f:
            for r in csv.DictReader(f):
                loc_id = r['id'].strip()
                name = r['name'].strip()
                state = normalize_state_name(r['state'])
                key = (name.lower(), state.lower())
                loc_obj = {
                    'id': loc_id,
                    'name': name,
                    'state': state,
                    'country': r.get('country', 'India').strip(),
                    'currency_code': r.get('currency_code', 'INR').strip(),
                    'description': r.get('description', '').strip(),
                    'latitude': float(r['latitude']) if r.get('latitude') else None,
                    'longitude': float(r['longitude']) if r.get('longitude') else None,
                    'region': 'Southern',
                    'created_at': r.get('created_at', '2026-09-04 00:00:00'),
                    'updated_at': r.get('updated_at', '2026-09-04 00:00:00')
                }
                locations[key] = loc_obj
                loc_id_map[loc_id] = loc_obj

    # 3. Load Excel sheets to discover all other locations
    xlsx_path = os.path.join(DATASETS_DIR, 'india_tourism_dataset-8.xlsx')
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        # Scan Dataset sheet
        if 'Dataset' in wb.sheetnames:
            for r in list(wb['Dataset'].iter_rows(values_only=True))[1:]:
                st = normalize_state_name(r[1])
                dist = str(r[2] or '').strip()
                if st and dist:
                    key = (dist.lower(), st.lower())
                    if key not in locations:
                        st_code = STATE_CODE_MAP.get(st.lower(), slugify(st)[:3])
                        loc_id = f"loc-{st_code}-{slugify(dist, 10)}"
                        counter = 1
                        orig_id = loc_id
                        while loc_id in loc_id_map:
                            loc_id = f"{orig_id}-{counter}"
                            counter += 1

                        region = STATE_REGION_MAP.get(st.lower(), 'Northern')
                        loc_obj = {
                            'id': loc_id,
                            'name': dist,
                            'state': st,
                            'country': 'India',
                            'currency_code': 'INR',
                            'description': f"Scenic and historic tourist district in {st}, known for {r[0]} and regional cultural attractions.",
                            'latitude': float(r[4]) if r[4] is not None else None,
                            'longitude': float(r[5]) if r[5] is not None else None,
                            'region': region,
                            'created_at': '2026-09-04 00:00:00',
                            'updated_at': '2026-09-04 00:00:00'
                        }
                        locations[key] = loc_obj
                        loc_id_map[loc_id] = loc_obj

        # Scan State_District_Spots sheet
        if 'State_District_Spots' in wb.sheetnames:
            for r in list(wb['State_District_Spots'].iter_rows(values_only=True))[1:]:
                st = normalize_state_name(r[0])
                dist = str(r[1] or '').strip()
                if st and dist:
                    key = (dist.lower(), st.lower())
                    if key not in locations:
                        st_code = STATE_CODE_MAP.get(st.lower(), slugify(st)[:3])
                        loc_id = f"loc-{st_code}-{slugify(dist.split('(')[0], 10)}"
                        counter = 1
                        orig_id = loc_id
                        while loc_id in loc_id_map:
                            loc_id = f"{orig_id}-{counter}"
                            counter += 1

                        region = STATE_REGION_MAP.get(st.lower(), 'Northern')
                        loc_obj = {
                            'id': loc_id,
                            'name': dist,
                            'state': st,
                            'country': 'India',
                            'currency_code': 'INR',
                            'description': f"Major travel destination and tourist district in {st}, hosting prominent attractions like {r[2]}.",
                            'latitude': None,
                            'longitude': None,
                            'region': region,
                            'created_at': '2026-09-04 00:00:00',
                            'updated_at': '2026-09-04 00:00:00'
                        }
                        locations[key] = loc_obj
                        loc_id_map[loc_id] = loc_obj

        # Scan Uploaded_Places_Extended sheet
        if 'Uploaded_Places_Extended' in wb.sheetnames:
            for r in list(wb['Uploaded_Places_Extended'].iter_rows(values_only=True))[1:]:
                st = normalize_state_name(r[0])
                city = str(r[1] or '').strip()
                if st and city:
                    key = (city.lower(), st.lower())
                    if key not in locations:
                        st_code = STATE_CODE_MAP.get(st.lower(), slugify(st)[:3])
                        loc_id = f"loc-{st_code}-{slugify(city, 10)}"
                        counter = 1
                        orig_id = loc_id
                        while loc_id in loc_id_map:
                            loc_id = f"{orig_id}-{counter}"
                            counter += 1

                        region = STATE_REGION_MAP.get(st.lower(), 'Northern')
                        loc_obj = {
                            'id': loc_id,
                            'name': city,
                            'state': st,
                            'country': 'India',
                            'currency_code': 'INR',
                            'description': f"Vibrant tourism destination in {st}, featuring celebrated landmarks including {r[2]}.",
                            'latitude': float(r[4]) if r[4] is not None else None,
                            'longitude': float(r[5]) if r[5] is not None else None,
                            'region': region,
                            'created_at': '2026-09-04 00:00:00',
                            'updated_at': '2026-09-04 00:00:00'
                        }
                        locations[key] = loc_obj
                        loc_id_map[loc_id] = loc_obj

    return locations, loc_id_map

def find_location_id(district_name, state_name, locations_dict, loc_id_map):
    st_norm = normalize_state_name(state_name).lower()
    dist_norm = str(district_name or '').strip().lower()

    if (dist_norm, st_norm) in locations_dict:
        return locations_dict[(dist_norm, st_norm)]['id']

    for (d, s), obj in locations_dict.items():
        if s == st_norm:
            if dist_norm in d or d in dist_norm:
                return obj['id']
            d_clean = re.sub(r'\(.*?\)', '', d).strip()
            dist_clean = re.sub(r'\(.*?\)', '', dist_norm).strip()
            if d_clean and dist_clean and (d_clean in dist_clean or dist_clean in d_clean):
                return obj['id']

    for (d, s), obj in locations_dict.items():
        if s == st_norm:
            return obj['id']

    return 'loc-chn'

def load_places(locations_dict, loc_id_map):
    places = {}
    place_ids = set()

    def generate_place_id(pname):
        base = f"plc-{slugify(pname, 32)}"
        pid = base
        c = 1
        while pid in place_ids:
            pid = f"{base}-{c}"
            c += 1
        place_ids.add(pid)
        return pid

    # 1. Load places_1.csv and places_2.csv
    for fname in ['places_1.csv', 'places_2.csv']:
        fpath = os.path.join(DATASETS_DIR, fname)
        if os.path.exists(fpath):
            with open(fpath, 'r', encoding='utf-8', errors='ignore') as f:
                for r in csv.DictReader(f):
                    name = r['name'].strip()
                    pid = r['id'].strip()
                    place_ids.add(pid)
                    norm_key = name.lower()

                    place_obj = {
                        'id': pid,
                        'location_id': r['location_id'].strip(),
                        'name': name,
                        'category': normalize_category(r.get('category', ''), name),
                        'avg_rating': parse_rating(r.get('avg_rating', 4.5)),
                        'review_count': int(float(r.get('review_count', 100) or 100)),
                        'entry_fee': parse_entry_fee(r.get('entry_fee', 0)),
                        'opening_hours': r.get('opening_hours', '09:00 - 18:00').strip(),
                        'latitude': float(r['latitude']) if r.get('latitude') else None,
                        'longitude': float(r['longitude']) if r.get('longitude') else None,
                        'description': r.get('description', '').strip() or f"{name} is a renowned attraction, welcoming travelers from around the world.",
                        'best_season': r.get('best_season', 'Oct-Mar').strip(),
                        'avg_visit_time': r.get('avg_visit_time', '1-2 hours').strip(),
                        'transport': r.get('transport', 'Accessible via local bus, auto, and taxi.').strip(),
                        'nearby_hotels': r.get('nearby_hotels', 'Budget to premium hotels and homestays available in vicinity.').strip(),
                        'nearby_restaurants': r.get('nearby_restaurants', 'Local authentic dining and multi-cuisine eateries nearby.').strip(),
                        'created_at': r.get('created_at', '2026-09-04 00:00:00'),
                        'updated_at': r.get('updated_at', '2026-09-04 00:00:00')
                    }
                    places[norm_key] = place_obj

    print(f"✅ Loaded {len(places)} places from places_1.csv & places_2.csv")

    # 2. Load Excel Workbook
    xlsx_path = os.path.join(DATASETS_DIR, 'india_tourism_dataset-8.xlsx')
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        # 2a. Sheet: Dataset (57 destinations across India)
        if 'Dataset' in wb.sheetnames:
            ds_sheet = wb['Dataset']
            for r in list(ds_sheet.iter_rows(values_only=True))[1:]:
                pname = str(r[0] or '').strip()
                if not pname: continue
                st = normalize_state_name(r[1])
                dist = str(r[2] or '').strip()
                norm_key = pname.lower()

                loc_id = find_location_id(dist, st, locations_dict, loc_id_map)
                lat = float(r[4]) if r[4] is not None else None
                lng = float(r[5]) if r[5] is not None else None
                desc = str(r[6] or '').strip()
                season = str(r[7] or 'Oct-Mar').strip()
                fee = parse_entry_fee(r[8])
                rating = parse_rating(r[9])
                visit_time = str(r[10] or '1-2 hours').strip()
                hotels = str(r[11] or 'Budget to luxury accommodations available nearby.').strip()
                dining = str(r[12] or 'Local regional cuisine and restaurants in vicinity.').strip()
                transport = str(r[13] or 'Well connected by road, bus, and taxi services.').strip()

                if norm_key in places:
                    p = places[norm_key]
                    if not p['latitude'] and lat: p['latitude'] = lat
                    if not p['longitude'] and lng: p['longitude'] = lng
                    if desc and len(desc) > len(p['description']): p['description'] = desc
                    if hotels: p['nearby_hotels'] = hotels
                    if dining: p['nearby_restaurants'] = dining
                    if transport: p['transport'] = transport
                else:
                    pid = generate_place_id(pname)
                    places[norm_key] = {
                        'id': pid,
                        'location_id': loc_id,
                        'name': pname,
                        'category': normalize_category(r[3], pname),
                        'avg_rating': rating,
                        'review_count': 1500,
                        'entry_fee': fee,
                        'opening_hours': '09:00 - 18:00',
                        'latitude': lat,
                        'longitude': lng,
                        'description': desc or f"{pname} is a prominent tourist landmark located in {dist}, {st}.",
                        'best_season': season,
                        'avg_visit_time': visit_time,
                        'transport': transport,
                        'nearby_hotels': hotels,
                        'nearby_restaurants': dining,
                        'created_at': '2026-09-04 00:00:00',
                        'updated_at': '2026-09-04 00:00:00'
                    }

        print(f"✅ Total places after 'Dataset' sheet: {len(places)}")

        # 2b. Sheet: State_District_Spots (235 spots)
        if 'State_District_Spots' in wb.sheetnames:
            sds_sheet = wb['State_District_Spots']
            for r in list(sds_sheet.iter_rows(values_only=True))[1:]:
                st = normalize_state_name(r[0])
                dist = str(r[1] or '').strip()
                pname = str(r[2] or '').strip()
                if not pname: continue
                norm_key = pname.lower()

                loc_id = find_location_id(dist, st, locations_dict, loc_id_map)
                cat = normalize_category(r[3], pname)
                season = str(r[4] or 'Oct-Mar').strip()
                fee = parse_entry_fee(r[5])
                rating = parse_rating(r[6])
                visit_time = str(r[7] or '1-2 hours').strip()
                hotels = str(r[9] or 'Budget to mid-range stays available.').strip()
                dining = str(r[10] or 'Regional eateries and cafes nearby.').strip() if len(r) > 10 else 'Regional eateries nearby.'

                if norm_key in places:
                    p = places[norm_key]
                    if not p.get('best_season') and season: p['best_season'] = season
                    if not p.get('avg_visit_time') and visit_time: p['avg_visit_time'] = visit_time
                else:
                    pid = generate_place_id(pname)
                    places[norm_key] = {
                        'id': pid,
                        'location_id': loc_id,
                        'name': pname,
                        'category': cat,
                        'avg_rating': rating,
                        'review_count': 950,
                        'entry_fee': fee,
                        'opening_hours': '08:30 - 18:30',
                        'latitude': None,
                        'longitude': None,
                        'description': f"{pname} is a celebrated {cat} attraction situated in {dist}, {st}, popular for sightseers and culture lovers.",
                        'best_season': season,
                        'avg_visit_time': visit_time,
                        'transport': f"Accessible by local buses, autos, and private vehicles from central {dist}.",
                        'nearby_hotels': hotels,
                        'nearby_restaurants': dining,
                        'created_at': '2026-09-04 00:00:00',
                        'updated_at': '2026-09-04 00:00:00'
                    }

        print(f"✅ Total places after 'State_District_Spots': {len(places)}")

        # 2c. Sheet: Uploaded_Places_Extended (1,019 spots)
        if 'Uploaded_Places_Extended' in wb.sheetnames:
            upe_sheet = wb['Uploaded_Places_Extended']
            for r in list(upe_sheet.iter_rows(values_only=True))[1:]:
                st = normalize_state_name(r[0])
                city = str(r[1] or '').strip()
                pname = str(r[2] or '').strip()
                if not pname: continue
                norm_key = pname.lower()

                lat = float(r[4]) if r[4] is not None else None
                lng = float(r[5]) if r[5] is not None else None
                rating = parse_rating(r[6])
                fee = parse_entry_fee(r[7])

                if norm_key in places:
                    p = places[norm_key]
                    if not p['latitude'] and lat: p['latitude'] = lat
                    if not p['longitude'] and lng: p['longitude'] = lng
                    if fee > 0 and p['entry_fee'] == 0: p['entry_fee'] = fee
                else:
                    loc_id = find_location_id(city, st, locations_dict, loc_id_map)
                    cat = normalize_category(r[3], pname)
                    pid = generate_place_id(pname)
                    places[norm_key] = {
                        'id': pid,
                        'location_id': loc_id,
                        'name': pname,
                        'category': cat,
                        'avg_rating': rating,
                        'review_count': 1200,
                        'entry_fee': fee,
                        'opening_hours': '09:00 - 18:00',
                        'latitude': lat,
                        'longitude': lng,
                        'description': f"{pname} is a premier {cat} destination in {city}, {st}, attracting tourists with its iconic significance.",
                        'best_season': 'Oct-Mar',
                        'avg_visit_time': '1-2 hours',
                        'transport': f"Well connected by city bus lines, cabs, and metro/rail routes in {city}.",
                        'nearby_hotels': 'Top rated hotels, boutique lodges, and homestays nearby.',
                        'nearby_restaurants': 'Variety of authentic multi-cuisine restaurants and street food in area.',
                        'created_at': '2026-09-04 00:00:00',
                        'updated_at': '2026-09-04 00:00:00'
                    }

        print(f"✅ Total places after 'Uploaded_Places_Extended': {len(places)}")

        # 2d. Sheet: Uploaded_Top_Places_Kaggle (325 spots)
        if 'Uploaded_Top_Places_Kaggle' in wb.sheetnames:
            k_sheet = wb['Uploaded_Top_Places_Kaggle']
            for r in list(k_sheet.iter_rows(values_only=True))[1:]:
                pname = str(r[3] or '').strip()
                norm_key = pname.lower()
                if norm_key in places:
                    p = places[norm_key]
                    if r[6] is not None and not p.get('avg_visit_time'):
                        p['avg_visit_time'] = f"{r[6]} hours"

    # Precise Google Maps coordinates for Chennai places & landmark fixes
    chennai_overrides = {
        # --- From places_1.csv & verified landmarks ---
        "plc-san-thome":              (13.03361, 80.27778),   # San Thome Basilica
        "plc-elliots-beach":          (12.99953, 80.27241),   # Edward Elliot's Beach
        "plc-fort-st-george":         (13.07972, 80.28694),   # Fort St. George & Museum
        "plc-egmore-museum":          (13.0706,  80.2567),    # Government Museum Chennai
        "plc-guindy-park":            (13.0089,  80.2406),    # Guindy National Park
        "plc-vandalur-zoo":           (12.87917, 80.08167),   # Arignar Anna Zoological Park
        "plc-dakshinachitra":         (12.82242, 80.2431),    # DakshinaChitra Heritage Museum
        "plc-dakshinachitra-museum":  (12.82242, 80.2431),    # DakshinaChitra Museum (exact same spot)
        "plc-valluvar-kottam":        (13.05441, 80.24175),   # Valluvar Kottam
        "plc-parthasarathy":          (13.0506,  80.2739),    # Arulmigu Parthasarathy Temple
        "plc-ashtalakshmi":           (12.9931,  80.2686),    # Ashtalakshmi Temple
        "plc-marundeeswarar":         (12.98556, 80.26139),   # Marundeeswarar Temple
        "plc-kalakshetra":            (12.9881,  80.265),     # Kalakshetra Foundation
        "plc-birla-planetarium":      (13.012,   80.2437),    # Birla Planetarium
        "plc-theosophical":           (13.0036,  80.2581),    # Theosophical Society Gardens
        "plc-semmozhi-poonga":        (13.0505,  80.2505),    # Semmozhi Poonga (Cathedral Rd)
        "plc-vadapalani-murugan":     (13.0528,  80.2136),    # Vadapalani Murugan Temple
        "plc-ripon-building":         (13.0825,  80.2715),    # Ripon Building & Victoria Public Hall
        "plc-st-thomas-mount":        (12.9972,  80.1639),    # St. Thomas Mount National Shrine
        "plc-rail-museum":            (13.0897,  80.2606),    # Chennai Rail Museum
        "plc-connemara-library":      (13.0706,  80.2567),    # Connemara Public Library (Egmore Museum complex)
        "plc-covelong-beach":         (12.7896,  80.2542),    # Covelong (Kovalam) Beach
        "plc-croc-bank":              (12.78,    80.239),     # Madras Crocodile Bank Trust
        "plc-vivekananda-house":      (13.0447,  80.2789),    # Vivekananda House (Illam)
        "plc-muttukadu":              (12.8227,  80.2419),    # Muttukadu Boat House
        "plc-water-activities-at-muttukadu": (12.8227, 80.2419), # Muttukadu Water Activities (same spot)
        "plc-cholamandal":            (12.92222, 80.25194),   # Cholamandal Artists' Village
        "plc-chetpet-eco-park":       (13.07412, 80.24238),   # Chetpet Eco Park
        "plc-thiruvanmiyur-beach":    (12.9736,  80.2665),    # Thiruvanmiyur Beach
        "plc-kalikambal":             (13.0872,  80.2889),    # Kalikambal Temple George Town
        "plc-thousand-lights":        (13.0547,  80.2422),    # Thousand Lights Mosque
        "plc-anna-nagar-tower":       (13.08678, 80.21435),   # Anna Nagar Tower Park

        # --- From Excel dataset & constant place alignments ---
        "plc-marina-beach":           (13.0500,  80.2824),    # Marina Beach
        "plc-barracuda-bayfishing":   (13.0500,  80.2824),    # Barracuda Bay Fishing (off Marina Beach - same spot)
        "plc-national-art-gallery":   (13.0706,  80.2567),    # National Art Gallery (inside Egmore Museum complex - same spot)
        "plc-pondy-bazaar":           (13.0399,  80.2388),    # Pondy Bazaar T. Nagar
        "plc-pulicat-lake":           (13.4167,  80.3167),    # Pulicat Lake
        "plc-mahabalipuram":          (12.6168,  80.1993),    # Mahabalipuram Shore Temple
        "plc-mgr-film-city":          (12.9866,  80.2492),    # MGR Film City Taramani
        "plc-adventure-sports-at-covelong-bea": (12.7896, 80.2542),  # Covelong Adventure Sports (same spot as Covelong Beach)
        "plc-scuba-diving-at-covelong-beach":   (12.7896, 80.2542),  # Covelong Scuba Diving (same spot as Covelong Beach)
        "plc-queensland-amusement-park":        (13.0450, 80.0105),  # Queensland Amusement Park (Palanjur NH4)
        "plc-ubbalamadugu-falls":     (13.6042,  79.9711),    # Ubbalamadugu Falls (Tada Falls)
        "plc-eachanari":              (10.9419,  76.9697),    # Eachanari Vinayagar Temple
    }
    for p in places.values():
        if p['id'] in chennai_overrides:
            p['latitude'], p['longitude'] = chennai_overrides[p['id']]
        if p.get('latitude') is not None and p.get('longitude') is not None:
            p['map_url'] = f"https://www.google.com/maps/search/?api=1&query={p['latitude']},{p['longitude']}"
        else:
            p['map_url'] = None

    return list(places.values())

def update_sqlite_db(locations_list, places_list):
    print(f"\n📁 Updating SQLite database at: {DB_PATH}...")
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("DROP TABLE IF EXISTS places;")
    cur.execute("DROP TABLE IF EXISTS locations;")

    cur.execute("""
    CREATE TABLE locations (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        state TEXT NOT NULL DEFAULT 'Tamil Nadu',
        country TEXT NOT NULL DEFAULT 'India',
        currency_code TEXT NOT NULL DEFAULT 'INR',
        description TEXT,
        latitude REAL,
        longitude REAL,
        region TEXT DEFAULT 'Southern',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    """)

    cur.execute("""
    CREATE TABLE places (
        id TEXT PRIMARY KEY,
        location_id TEXT NOT NULL,
        name TEXT NOT NULL,
        category TEXT NOT NULL,
        avg_rating REAL DEFAULT 0.00,
        review_count INTEGER DEFAULT 0,
        entry_fee REAL DEFAULT 0.00,
        opening_hours TEXT,
        latitude REAL,
        longitude REAL,
        map_url TEXT,
        description TEXT,
        best_season TEXT,
        avg_visit_time TEXT,
        transport TEXT,
        nearby_hotels TEXT,
        nearby_restaurants TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (location_id) REFERENCES locations (id) ON DELETE CASCADE
    );
    """)

    loc_tuples = [
        (
            l['id'], l['name'], l['state'], l['country'], l['currency_code'],
            l['description'], l['latitude'], l['longitude'], l['region'],
            l['created_at'], l['updated_at']
        )
        for l in locations_list
    ]
    cur.executemany("""
    INSERT INTO locations (id, name, state, country, currency_code, description, latitude, longitude, region, created_at, updated_at)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
    """, loc_tuples)

    place_tuples = [
        (
            p['id'], p['location_id'], p['name'], p['category'], p['avg_rating'],
            p['review_count'], p['entry_fee'], p['opening_hours'], p['latitude'],
            p['longitude'], p.get('map_url'), p['description'], p['best_season'], p['avg_visit_time'],
            p['transport'], p['nearby_hotels'], p['nearby_restaurants'],
            p['created_at'], p['updated_at']
        )
        for p in places_list
    ]
    cur.executemany("""
    INSERT INTO places (
        id, location_id, name, category, avg_rating, review_count, entry_fee,
        opening_hours, latitude, longitude, map_url, description, best_season, avg_visit_time,
        transport, nearby_hotels, nearby_restaurants, created_at, updated_at
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
    """, place_tuples)

    conn.commit()

    cur.execute("SELECT count(*) FROM locations;")
    loc_cnt = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM places;")
    plc_cnt = cur.fetchone()[0]
    conn.close()

    print(f"🎉 SQLite Successfully Updated: {loc_cnt} Locations, {plc_cnt} Places.")

def update_seed_sql(locations_list, places_list):
    print(f"\n📝 Updating MySQL seed file at: {SEED_SQL_PATH}...")

    def sql_escape(val):
        if val is None:
            return 'NULL'
        s = str(val).replace('\\', '\\\\').replace("'", "''")
        return f"'{s}'"

    def sql_num(val):
        if val is None:
            return 'NULL'
        return str(val)

    lines = [
        "-- =============================================================================",
        "-- TripNova Live Tourism Platform - Master Database Seed File",
        "-- Auto-generated by backend/scripts/syncDatasets.py",
        f"-- Date Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"-- Scale: {len(locations_list)} Destinations & {len(places_list)} Tourist Spots Across India",
        "-- =============================================================================",
        "",
        "SET @OLD_UNIQUE_CHECKS=@@UNIQUE_CHECKS, UNIQUE_CHECKS=0;",
        "SET @OLD_FOREIGN_KEY_CHECKS=@@FOREIGN_KEY_CHECKS, FOREIGN_KEY_CHECKS=0;",
        "SET NAMES utf8mb4;",
        "",
        "DELETE FROM `places`;",
        "DELETE FROM `locations`;",
        "",
        "-- -----------------------------------------------------------------------------",
        f"-- 1. Seed Locations ({len(locations_list)} Administrative Cities & Districts Across India)",
        "-- -----------------------------------------------------------------------------",
        "INSERT INTO `locations` (`id`, `name`, `state`, `country`, `currency_code`, `description`, `latitude`, `longitude`, `region`, `created_at`, `updated_at`) VALUES"
    ]

    loc_values = []
    for l in locations_list:
        loc_values.append(
            f"({sql_escape(l['id'])}, {sql_escape(l['name'])}, {sql_escape(l['state'])}, {sql_escape(l['country'])}, {sql_escape(l['currency_code'])}, {sql_escape(l['description'])}, {sql_num(l['latitude'])}, {sql_num(l['longitude'])}, {sql_escape(l['region'])}, {sql_escape(l['created_at'])}, {sql_escape(l['updated_at'])})"
        )
    lines.append(",\n".join(loc_values) + ";")
    lines.append("")

    lines.append("-- -----------------------------------------------------------------------------")
    lines.append(f"-- 2. Seed Places ({len(places_list)} Attractions Across India with Full Attributes & Exact Map URLs)")
    lines.append("-- -----------------------------------------------------------------------------")

    chunk_size = 100
    for chunk_idx in range(0, len(places_list), chunk_size):
        chunk = places_list[chunk_idx:chunk_idx + chunk_size]
        lines.append("INSERT INTO `places` (`id`, `location_id`, `name`, `category`, `avg_rating`, `review_count`, `entry_fee`, `opening_hours`, `latitude`, `longitude`, `map_url`, `description`, `best_season`, `avg_visit_time`, `transport`, `nearby_hotels`, `nearby_restaurants`, `created_at`, `updated_at`) VALUES")
        chunk_vals = []
        for p in chunk:
            chunk_vals.append(
                f"({sql_escape(p['id'])}, {sql_escape(p['location_id'])}, {sql_escape(p['name'])}, {sql_escape(p['category'])}, {sql_num(p['avg_rating'])}, {sql_num(p['review_count'])}, {sql_num(p['entry_fee'])}, {sql_escape(p['opening_hours'])}, {sql_num(p['latitude'])}, {sql_num(p['longitude'])}, {sql_escape(p.get('map_url'))}, {sql_escape(p['description'])}, {sql_escape(p['best_season'])}, {sql_escape(p['avg_visit_time'])}, {sql_escape(p['transport'])}, {sql_escape(p['nearby_hotels'])}, {sql_escape(p['nearby_restaurants'])}, {sql_escape(p['created_at'])}, {sql_escape(p['updated_at'])})"
            )
        lines.append(",\n".join(chunk_vals) + ";")
        lines.append("")

    lines.append("SET FOREIGN_KEY_CHECKS=@OLD_FOREIGN_KEY_CHECKS;")
    lines.append("SET UNIQUE_CHECKS=@OLD_UNIQUE_CHECKS;")
    lines.append("")

    with open(SEED_SQL_PATH, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))

    print(f"🎉 MySQL Seed File Generated: {SEED_SQL_PATH} ({os.path.getsize(SEED_SQL_PATH)} bytes)")

def main():
    print("🚀 Starting TripNova Dataset Synchronization...")
    locations_dict, loc_id_map = load_locations()
    print(f"✅ Loaded {len(locations_dict)} unique locations across all Indian States & UTs.")

    places_list = load_places(locations_dict, loc_id_map)
    print(f"✅ Loaded {len(places_list)} unified tourist attractions.")

    locations_list = list(locations_dict.values())
    locations_list.sort(key=lambda x: (0 if x['state'] == 'Tamil Nadu' else 1, x['state'], x['name']))

    update_sqlite_db(locations_list, places_list)
    update_seed_sql(locations_list, places_list)

    print("\n========================================================")
    print("📊 DATASET SYNCHRONIZATION COMPLETE SUMMARY:")
    print(f"📍 Total Locations: {len(locations_list)}")
    print(f"🏛️ Total Places: {len(places_list)}")
    states = set(l['state'] for l in locations_list)
    print(f"🌐 States & UTs Covered: {len(states)}")
    cats = {}
    for p in places_list:
        cats[p['category']] = cats.get(p['category'], 0) + 1
    print(f"📂 Category Breakdown: {cats}")
    print("========================================================")

if __name__ == '__main__':
    main()
